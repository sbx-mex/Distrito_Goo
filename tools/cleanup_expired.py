#!/usr/bin/env python3
"""Audita o retira filas vencidas del CMS con confirmación explícita y trazabilidad."""
from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

CONFIRMATION = "RETIRAR_CONTENIDO_EXPIRADO"
SHEETS = {
    "Eventos": ("Fecha Fin", "ID", "Nombre Evento"),
    "Informativo": ("Vigencia Fin", "ID", "Actividad"),
}


def as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def audit(workbook: Path, reference: date, grace_days: int) -> tuple[dict[str, object], dict[str, list[int]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    cutoff = reference - timedelta(days=max(0, grace_days))
    candidates: list[dict[str, object]] = []
    rows_by_sheet: dict[str, list[int]] = {}
    for sheet_name, (end_header, id_header, title_header) in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows())
        if not sheet_rows:
            continue
        headers = {str(cell.value or "").strip(): index for index, cell in enumerate(sheet_rows[0])}
        if end_header not in headers:
            continue
        for row_number, cells in enumerate(sheet_rows[1:], start=2):
            end = as_date(cells[headers[end_header]].value)
            if end is None or end >= cutoff:
                continue
            item = {
                "sheet": sheet_name,
                "row": row_number,
                "id": str(cells[headers.get(id_header, 0)].value or ""),
                "title": str(cells[headers.get(title_header, 1)].value or ""),
                "end": end.isoformat(),
                "reason": f"vigencia final anterior a {cutoff.isoformat()} (gracia {grace_days} días)",
            }
            candidates.append(item)
            rows_by_sheet.setdefault(sheet_name, []).append(row_number)
    wb.close()
    return {
        "ok": True,
        "asOf": reference.isoformat(),
        "graceDays": grace_days,
        "cutoff": cutoff.isoformat(),
        "candidates": candidates,
        "summary": {"candidates": len(candidates), "removed": 0},
    }, rows_by_sheet


def apply(workbook: Path, rows_by_sheet: dict[str, list[int]]) -> int:
    wb = load_workbook(workbook)
    removed = 0
    for sheet_name, rows in rows_by_sheet.items():
        ws = wb[sheet_name]
        for row in sorted(rows, reverse=True):
            ws.delete_rows(row, 1)
            removed += 1
        for table in ws.tables.values():
            start = table.ref.split(":", 1)[0]
            end_col = table.ref.split(":", 1)[1].rstrip("0123456789")
            table.ref = f"{start}:{end_col}{ws.max_row}"
    wb.save(workbook)
    return removed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("cms", type=Path)
    parser.add_argument("--as-of", default=date.today().isoformat())
    parser.add_argument("--grace-days", type=int, default=7)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm", default="")
    parser.add_argument("--report", type=Path, default=Path("reports/expired-content.json"))
    args = parser.parse_args()
    reference = date.fromisoformat(args.as_of)
    report, rows = audit(args.cms, reference, args.grace_days)
    if args.apply:
        if args.confirm != CONFIRMATION:
            raise SystemExit(f"Confirmación inválida. Usa --confirm {CONFIRMATION}")
        report["summary"]["removed"] = apply(args.cms, rows)
        report["mode"] = "apply"
    else:
        report["mode"] = "audit"
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
