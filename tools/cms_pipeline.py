#!/usr/bin/env python3
"""Valida Distrito_Go_CMS.xlsx y genera los JSON estáticos de Distrito Goo.

No se usa en el navegador ni requiere backend. Está pensado para ejecución local o CI.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from datetime import date, datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_HEADERS = {
    "Informativo": ["ID", "Actividad", "Descripción", "Link /Imagen", "Frecuencia", "Prioridad", "Categoría", "Icono", "Color", "Visible"],
    "WFM": ["Regla WFM"],
    "BT": ["Mes", "SBX", "NO. EMPLEADO", "NOMBRE COMPLETO", "CECO", "TIENDA", "REGION", "DM", "HRBP", "RD", "GB180", "ANTIGÜEDAD", "JORNADA", "ESTATUS ALTA"],
    "TBW": ["Corte", "SBX", "NOMBRE", "CeCo", "TIENDA", "PUESTO", "Región", "DM", "HRBP", "Fecha de ingreso", "Días de antigüedad", "To be Welcoming fundacional \nDía 36 al 60", "To be Welcoming Sesgo de edad\nDía 60 al 90", "To be Welcoming Discapacidad \nDía 60 al 90", "To be Welcoming Género\nDía 60 al 90", "To be Welcoming Sexualidad\ndía 90- 120", "Avance"],
    "SS": ["Mes", "SBX", "NO. EMPLEADO", "NOMBRE COMPLETO", "CECO", "TIENDA", "REGION", "DM", "HRBP", "RD", "mes de solicitud", "BT", "ESTATUS ALTA"],
    "Links": ["Categoria", "Grupo", "Icono", "Nombre", "Tipo", "URL", "WebURL", "Package", "PlayStore", "Notas", "Favorito", "Orden"],
    "Eventos": ["ID", "Nombre Evento", "Descripción", "Fecha Inicio", "Fecha Fin", "Región", "Distrito", "Tienda", "Publicar", "Link/Imagen", "Imagen", "Tipo de acción"],
    "Actividades_Semanales": ["ID", "Actividad", "Descripción", "Día", "Hora / Corte", "Icono", "Color", "Link"],
    "Actividades_Diaria": ["ID", "Actividad", "Descripción", "Link / Imagen", "Frecuencia", "Prioridad", "Categoría", "Icono", "Color", "Visible"],
    "Duty_Roster": ["Orden", "Día", "Estaciones", "Imágenes", "Color", "Enfoque"],
    "Duty_Detail": ["Día", "Estación", "Categoría", "Orden", "Actividad", "Icono", "Crítico"],
    "Checklist_Apertura": ["Actividad", "Orden", "Concepto", "Icono"],
    "Identidad": ["Identificador", "Sección", "Campo", "Valor", "Color", "Estilo", "Visible", "Notas"],
    "Aniversarios_Cumpleanos": ["ID", "Tipo", "Nombre Partner", "Fecha", "Puesto", "Tienda", "Distrito", "Región", "Publicar", "NUM_EMP", "CECO"],
}

OPTIONAL_HEADERS = {
    "Informativo": ["Etiqueta", "Vigencia Inicio", "Vigencia Fin", "Orden", "Mostrar Inicio", "Mostrar Explorar", "Acceso Rápido", "CTA", "Link Detalle", "Recurso Secundario", "Puntos Clave", "Checklist Evaluación"],
}

def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().casefold()
    return re.sub(r"[^a-z0-9]+", "", text)

def canonical_header_map(sheet_name: str) -> dict[str, str]:
    names = [*REQUIRED_HEADERS.get(sheet_name, []), *OPTIONAL_HEADERS.get(sheet_name, [])]
    return {normalize_header(name): name for name in names}

CATEGORY_META = {
    "operacion": ("Operación", "⚙️", "#006241"),
    "personas": ("Personas", "👥", "#00754A"),
    "resultados": ("Resultados", "📊", "#1E6A8D"),
    "soporte": ("Soporte", "🛟", "#8A5A00"),
    "aprendizaje": ("Aprendizaje", "🎓", "#6B4EA0"),
    "otros": ("Otros", "🔗", "#5F6368"),
}


def plain(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def truthy(value: Any) -> bool:
    return str(value).strip().casefold() in {"true", "verdadero", "si", "sí", "1", "yes"}


def slug(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().casefold()
    return re.sub(r"[^a-z0-9]+", "-", text).strip("-") or "item"


def short(text: Any, limit: int = 118) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    return value if len(value) <= limit else value[: limit - 1].rstrip() + "…"


def read_sheet(ws, sheet_name: str = "") -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    canonical = canonical_header_map(sheet_name)
    headers = []
    for value in rows[0]:
        raw = str(value).strip() if value is not None else ""
        headers.append(canonical.get(normalize_header(raw), raw))
    output = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(v not in (None, "") for v in row):
            continue
        record = {headers[i]: plain(row[i]) if i < len(row) else "" for i in range(len(headers)) if headers[i]}
        record["__row__"] = row_number
        output.append(record)
    return output


def apply_header_overlays(root: Path, sheets: dict[str, list[dict[str, Any]]]) -> dict[str, list[dict[str, Any]]]:
    """Fusiona cms-overlays/<Pestaña>.csv por encabezados y llave primaria, sin depender del orden de columnas."""
    overlay_dir = root / "cms-overlays"
    if not overlay_dir.is_dir():
        return sheets
    primary_keys = {"Informativo": "ID", "Eventos": "ID", "Links": "Nombre", "Actividades_Semanales": "ID", "Actividades_Diaria": "ID"}
    for sheet_name, base_rows in sheets.items():
        path = overlay_dir / f"{sheet_name}.csv"
        if not path.is_file():
            continue
        canonical = canonical_header_map(sheet_name)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            overlay_rows = []
            for number, source in enumerate(reader, start=2):
                record = {}
                for header, value in source.items():
                    raw = str(header or "").strip()
                    name = canonical.get(normalize_header(raw), raw)
                    if name:
                        record[name] = plain(value)
                if any(str(value or "").strip() for value in record.values()):
                    record["__row__"] = number
                    overlay_rows.append(record)
        key = primary_keys.get(sheet_name)
        if not key:
            sheets[sheet_name].extend(overlay_rows)
            continue
        positions = {str(row.get(key, "")).strip().casefold(): idx for idx, row in enumerate(base_rows) if str(row.get(key, "")).strip()}
        for record in overlay_rows:
            marker = str(record.get(key, "")).strip().casefold()
            if marker and marker in positions:
                merged = {**base_rows[positions[marker]], **record}
                base_rows[positions[marker]] = merged
            else:
                base_rows.append(record)
    return sheets


def validate_workbook(path: Path) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    errors: list[str] = []
    sheets: dict[str, list[dict[str, Any]]] = {}
    for name, required in REQUIRED_HEADERS.items():
        if name not in wb.sheetnames:
            errors.append(f"Falta la pestaña obligatoria: {name}")
            continue
        ws = wb[name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        canonical = canonical_header_map(name)
        actual = [canonical.get(normalize_header(v), str(v).strip() if v is not None else "") for v in first]
        actual_keys = {normalize_header(v) for v in actual if v}
        missing = [h for h in required if normalize_header(h) not in actual_keys]
        if missing:
            errors.append(f"{name}: faltan encabezados: {', '.join(missing)}")
        sheets[name] = read_sheet(ws, name)
    return sheets, errors


NETWORK_FIRST_IMAGE = "resumen_comunicado_semana_actual.png"


@lru_cache(maxsize=256)
def asset_version(path: Path) -> str:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()[:12]
    return f"?v={digest}"


@lru_cache(maxsize=8)
def asset_index(root: Path, folder: str) -> dict[str, Path]:
    """Indexa cada carpeta una sola vez durante la compilación."""
    base = root / folder
    return {path.name.casefold(): path for path in base.iterdir() if path.is_file()} if base.is_dir() else {}


def image_path(value: Any, root: Path, folder: str = "assets/photos") -> str:
    name = Path(str(value or "").strip()).name
    if not name:
        return ""
    matches = asset_index(root, folder)
    match = matches.get(name.casefold())
    if not match:
        return f"{folder}/{name}"

    # El resumen semanal conserva su nombre PNG para mantener network-first.
    # El resto prefiere el WebP generado por el pipeline, sin invalidar el
    # nombre original indicado en el CMS.
    preferred = match
    if match.name.casefold() != NETWORK_FIRST_IMAGE.casefold():
        optimized = matches.get(f"{match.stem}.webp".casefold())
        if optimized:
            preferred = optimized
    return preferred.relative_to(root).as_posix() + asset_version(preferred)

def original_image_path(value: Any, root: Path, folder: str = "assets/photos") -> str:
    name = Path(str(value or "").strip().split("?", 1)[0].split("#", 1)[0]).name
    if not name:
        return ""
    matches = asset_index(root, folder)
    match = matches.get(name.casefold())
    if not match:
        return f"{folder}/{name}"
    return match.relative_to(root).as_posix() + asset_version(match)

def thumbnail_image_path(value: Any, root: Path, folder: str = "assets/photos") -> str:
    name = Path(str(value or "").strip().split("?", 1)[0].split("#", 1)[0]).name
    if not name:
        return ""
    matches = asset_index(root, folder)
    source = matches.get(name.casefold())
    if not source:
        return image_path(value, root, folder)
    thumbnail = matches.get(f"{source.stem}.thumb.webp".casefold())
    chosen = thumbnail or source
    return chosen.relative_to(root).as_posix() + asset_version(chosen)

def resource(value: Any, root: Path) -> tuple[str, str, str]:
    text = str(value or "").strip()
    if not text:
        return "", "", ""
    if re.match(r"^https?://", text, re.I):
        return text, "link", ""
    return image_path(text, root), "imagen", original_image_path(text, root)


def valid_http_url(value: Any) -> str:
    """Acepta solo destinos completos; rechaza abreviaciones y marcadores."""
    text = str(value or "").strip()
    if not re.match(r"^https?://", text, re.I) or "..." in text or re.search(r"[{}<>]", text):
        return ""
    return text


def partner_key(row: dict[str, Any]) -> str:
    """Devuelve el identificador real más estable disponible en el CMS."""
    return str(row.get("SBX") or row.get("NO. EMPLEADO") or "").strip()


def clean_partner_text(value: Any) -> str:
    """Normaliza espacios del CMS sin alterar acentos, nombres o identificadores."""
    return re.sub(r"\s+", " ", str(value or "")).strip()


def group_partners_by_store(items: list[dict[str, Any]], route: str) -> list[dict[str, Any]]:
    """Agrupa la vista informativa por tienda para evitar metadatos repetidos."""
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        store = clean_partner_text(item.get("tienda")) or "Tienda no indicada"
        partner = {
            "id": item["id"],
            "nombre": clean_partner_text(item.get("nombre")),
        }
        if route == "alta":
            partner["programa"] = clean_partner_text(item.get("programa"))
        else:
            partner["avance"] = item.get("avance", "")
        grouped.setdefault(store, []).append(partner)
    return [
        {
            "tienda": store,
            "total": len(partners),
            "partners": sorted(partners, key=lambda item: item["nombre"].casefold()),
        }
        for store, partners in sorted(grouped.items(), key=lambda item: item[0].casefold())
    ]


def build_partner_development(
    bt: list[dict[str, Any]],
    ss: list[dict[str, Any]],
    tbw: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Normaliza únicamente los datos que consume la vista Desarrollo Partner."""
    courses_by_id: dict[str, dict[str, Any]] = {}
    pending_by_id: dict[str, dict[str, Any]] = {}
    discarded = {"cursosSinIdONombre": 0, "cursosDuplicados": 0, "tbwSinIdONombre": 0, "tbwNoPendientes": 0, "tbwDuplicados": 0}

    for program, rows in (("BT", bt), ("SS", ss)):
        for row in rows:
            item_id = partner_key(row)
            name = clean_partner_text(row.get("NOMBRE COMPLETO") or row.get("NOMBRE"))
            if not item_id or not name:
                discarded["cursosSinIdONombre"] += 1
                continue
            if item_id in courses_by_id:
                discarded["cursosDuplicados"] += 1
                existing = courses_by_id[item_id]
                if program not in existing["programa"].split(" / "):
                    existing["programa"] = f"{existing['programa']} / {program}"
                continue
            courses_by_id[item_id] = {
                "id": item_id,
                "nombre": name,
                "tienda": clean_partner_text(row.get("TIENDA")),
                "estatus": clean_partner_text(row.get("ESTATUS ALTA")),
                "programa": program,
                "avance": clean_partner_text(row.get("GB180") or row.get("BT")),
                "fecha": clean_partner_text(row.get("Mes") or row.get("mes de solicitud")),
            }

    pending_values = {"incompleto", "en curso"}
    for row in tbw:
        item_id = partner_key(row)
        name = clean_partner_text(row.get("NOMBRE") or row.get("NOMBRE COMPLETO"))
        if not item_id or not name:
            discarded["tbwSinIdONombre"] += 1
            continue
        statuses = [
            str(value).strip()
            for key, value in row.items()
            if str(key).startswith("To be Welcoming") and str(value).strip()
        ]
        explicit_pending = [value for value in statuses if value.casefold() in pending_values]
        if not explicit_pending:
            discarded["tbwNoPendientes"] += 1
            continue
        if item_id in pending_by_id:
            discarded["tbwDuplicados"] += 1
            continue
        raw_progress = row.get("Avance", "")
        try:
            numeric_progress = float(raw_progress)
            progress = round(numeric_progress * 100 if 0 <= numeric_progress <= 1 else numeric_progress)
        except (TypeError, ValueError):
            progress = ""
        pending_by_id[item_id] = {
            "id": item_id,
            "nombre": name,
            "tienda": clean_partner_text(row.get("TIENDA")),
            "estatus": "Pendiente",
            "avance": progress,
            "fecha": clean_partner_text(row.get("Corte")),
        }

    courses = sorted(courses_by_id.values(), key=lambda item: (item["tienda"].casefold(), item["nombre"].casefold()))
    pending = sorted(pending_by_id.values(), key=lambda item: (item["tienda"].casefold(), item["nombre"].casefold()))
    source_dates = [item["fecha"] for item in pending if re.match(r"^\d{4}-\d{2}-\d{2}$", item["fecha"])]
    updated = max(source_dates) if source_dates else date.today().isoformat()
    course_periods = sorted({item["fecha"] for item in courses if item["fecha"]}, key=str.casefold)
    course_programs = {
        program: sum(program in item["programa"].split(" / ") for item in courses)
        for program in ("BT", "SS")
    }
    pending_started = sum(
        isinstance(item.get("avance"), (int, float)) and item["avance"] > 0
        for item in pending
    )
    data = {
        "actualizado": updated,
        "vistas": {
            "alta": {
                "titulo": "Cursos de Alta",
                "periodo": " · ".join(course_periods),
                "total": len(courses),
                "tiendas": len({item["tienda"] for item in courses if item["tienda"]}),
                "programas": course_programs,
                "grupos": group_partners_by_store(courses, "alta"),
            },
            "tbw": {
                "titulo": "TBW",
                "corte": updated,
                "total": len(pending),
                "tiendas": len({item["tienda"] for item in pending if item["tienda"]}),
                "progreso": {
                    "sinIniciar": len(pending) - pending_started,
                    "enCurso": pending_started,
                },
                "grupos": group_partners_by_store(pending, "tbw"),
            },
        },
        "cursosAlta": courses,
        "tbwPendientes": pending,
    }
    report = {
        "ok": True,
        "actualizado": updated,
        "procesados": {"BT": len(bt), "SS": len(ss), "TBW": len(tbw)},
        "publicados": {"cursosAlta": len(courses), "tbwPendientes": len(pending)},
        "descartados": discarded,
    }
    return data, report


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def canonical_hash(value: Any) -> str:
    """Huella estable de datos normalizados, independiente del formato del XLSX."""
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def source_metadata(root: Path, sheets: dict[str, list[dict[str, Any]]], source_path: Path | None) -> dict[str, Any]:
    raw = {name: [{k: v for k, v in row.items() if k != "__row__"} for row in rows] for name, rows in sheets.items()}
    source_sha = hashlib.sha256(source_path.read_bytes()).hexdigest() if source_path and source_path.is_file() else canonical_hash(raw)
    previous_path = root / "data" / "cms-build.v1.json"
    previous = json.loads(previous_path.read_text(encoding="utf-8")) if previous_path.exists() else {}
    generated_at = previous.get("generatedAt") if previous.get("sourceSha256") == source_sha else None
    return {
        "schemaVersion": 2,
        "source": source_path.name if source_path else "Distrito_Go_CMS_v2_actualizado.xlsx",
        "sourceSha256": source_sha,
        "generatedAt": generated_at or datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "records": sum(len(rows) for rows in sheets.values()),
        "sheets": {
            name: {"records": len(rows), "contentSha256": canonical_hash(raw[name])}
            for name, rows in sheets.items()
        },
    }


def build(root: Path, sheets: dict[str, list[dict[str, Any]]], source_path: Path | None = None) -> list[Path]:
    data = root / "data"
    changed: list[Path] = []

    def emit(name: str, value: Any):
        target = data / name
        before = target.read_text(encoding="utf-8") if target.exists() else None
        text = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
        if text != before:
            target.write_text(text, encoding="utf-8")
            changed.append(target)

    def emit_report(name: str, value: Any):
        target = root / "reports" / name
        target.parent.mkdir(parents=True, exist_ok=True)
        before = target.read_text(encoding="utf-8") if target.exists() else None
        text = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
        if text != before:
            target.write_text(text, encoding="utf-8")
            changed.append(target)

    raw = {name: [{k: v for k, v in row.items() if k != "__row__"} for row in rows] for name, rows in sheets.items()}

    daily = []
    for row in raw["Actividades_Diaria"]:
        rec, typ, original = resource(row.pop("Link / Imagen", ""), root)
        row["ID"] = int(row["ID"])
        row["Prioridad"] = int(row["Prioridad"])
        row["Visible"] = truthy(row["Visible"])
        thumbnail = thumbnail_image_path(original or rec, root) if typ == "imagen" else ""
        row.update({"Recurso": rec, "MiniaturaRecurso": thumbnail, "OriginalRecurso": original, "TipoRecurso": typ, "DescripcionBreve": short(row.get("Descripción"))})
        daily.append(row)

    info = []
    for row in raw["Informativo"]:
        rec, typ, original = resource(row.pop("Link /Imagen", ""), root)
        row["ID"] = int(row["ID"])
        row["Prioridad"] = int(row["Prioridad"])
        if str(row.get("Orden", "")).strip():
            try:
                row["Orden"] = int(float(str(row["Orden"]).strip()))
            except ValueError:
                pass
        row["Visible"] = truthy(row["Visible"])
        for field in ("Mostrar Inicio", "Mostrar Explorar"):
            if field in row:
                row[field] = truthy(row[field])
        thumbnail = thumbnail_image_path(original or rec, root) if typ == "imagen" else ""
        detail_raw = str(row.get("Link Detalle", "") or "").strip()
        if re.match(r"^https?://", detail_raw, re.I):
            detail_link = valid_http_url(detail_raw)
        else:
            detail_name = Path(detail_raw).name if detail_raw else ""
            detail_path = root / "assets/docs" / detail_name if detail_name else None
            detail_link = detail_path.relative_to(root).as_posix() if detail_path and detail_path.is_file() else ""
        secondary_raw = str(row.get("Recurso Secundario", "") or "").strip()
        secondary = image_path(secondary_raw, root) if secondary_raw else ""
        secondary_original = original_image_path(secondary_raw, root) if secondary_raw else ""
        row.update({"Recurso": rec, "MiniaturaRecurso": thumbnail, "OriginalRecurso": original, "TipoRecurso": typ, "DescripcionBreve": short(row.get("Descripción")), "LinkDetalle": detail_link, "RecursoSecundario": secondary, "RecursoSecundarioOriginal": secondary_original})
        info.append(row)

    weekly = raw["Actividades_Semanales"]
    for row in weekly:
        row["ID"] = int(row["ID"])

    events = []
    for row in raw["Eventos"]:
        if not truthy(row.get("Publicar")):
            continue
        mixed = str(row.pop("Link/Imagen", "") or "").strip()
        action = str(row.pop("Tipo de acción", "") or "").strip().casefold()
        action_name = {"enlace": "Enlace", "imagen": "Imagen", "informativo": "Informativo"}.get(action, "Informativo")
        link = valid_http_url(mixed) if action_name == "Enlace" else ""
        image_field = str(row.get("Imagen", "") or "").strip()
        image_in_icon_field = image_field if re.search(r"\.(?:avif|gif|jpe?g|png|webp)$", image_field, re.I) else ""
        mixed_image = mixed if re.search(r"\.(?:avif|gif|jpe?g|png|webp)$", mixed, re.I) else ""
        img_candidate = image_in_icon_field or mixed_image
        icon = "🚨" if image_in_icon_field else image_field
        events.append({
            "ID": row.get("ID", ""), "Actividad": row.get("Nombre Evento", ""),
            "Contexto / Recordatorio": row.get("Descripción", ""),
            "Fecha Inicio": row.get("Fecha Inicio", ""), "Fecha Fin": row.get("Fecha Fin", ""),
            "Región": row.get("Región", ""), "Distrito": row.get("Distrito", ""),
            "Tienda": row.get("Tienda", ""), "Publicar": True, "Imagen": icon, "TipoAccion": action_name,
            "Link": link, "ImagenPath": image_path(img_candidate, root) if img_candidate else "",
            "MiniaturaPath": thumbnail_image_path(img_candidate, root) if img_candidate else "",
            "ImagenOriginal": original_image_path(img_candidate, root) if img_candidate else ""
        })

    duty_roster = []
    for row in raw["Duty_Roster"]:
        names = [x.strip() for x in str(row.get("Imágenes", "")).split(",") if x.strip()]
        paths = [image_path(x, root, "assets/premium/duty-roster") for x in names]
        thumbnails = [thumbnail_image_path(x, root, "assets/premium/duty-roster") for x in names]
        originals = [original_image_path(x, root, "assets/premium/duty-roster") for x in names]
        duty_roster.append({**row, "Orden": int(row["Orden"]), "ImagenesPath": paths, "MiniaturasPath": thumbnails, "ImagenesOriginales": originals, "ImagenOriginal": originals[0] if originals else "", "Premium": bool(paths), "Link": paths[0] if paths else ""})

    duty_detail = raw["Duty_Detail"]
    for row in duty_detail:
        row["Orden"] = int(row["Orden"])
        row["Crítico"] = truthy(row["Crítico"])

    checklist = raw["Checklist_Apertura"]
    for row in checklist:
        row["Orden"] = int(row["Orden"])

    links = raw["Links"]
    old_tools = json.loads((data / "herramientas.v10.json").read_text(encoding="utf-8")) if (data / "herramientas.v10.json").exists() else []
    old_tools_by_name = {str(t.get("nombre", "")).strip().casefold(): t for t in old_tools}
    old_categories = json.loads((data / "categorias.v10.json").read_text(encoding="utf-8")) if (data / "categorias.v10.json").exists() else []
    old_categories_by_id = {c.get("id"): c for c in old_categories}
    tools = []
    used_ids: set[str] = set()
    for row in links:
        previous = old_tools_by_name.get(str(row["Nombre"]).strip().casefold(), {})
        base_id = previous.get("id") or slug(row["Nombre"])
        item_id = base_id
        suffix = 2
        while item_id in used_ids:
            item_id = f"{base_id}-{suffix}"
            suffix += 1
        used_ids.add(item_id)
        cat_id = previous.get("categoriaId") or slug(row["Categoria"])
        if cat_id not in CATEGORY_META:
            CATEGORY_META.setdefault(cat_id, (str(row["Categoria"]), "🔗", "#5F6368"))
        keyword_source = " ".join(str(row.get(k, "")) for k in ("Categoria", "Grupo", "Nombre", "Tipo", "Notas"))
        keywords = sorted({slug(word) for word in re.findall(r"[\wÁÉÍÓÚÜÑáéíóúüñ]+", keyword_source) if len(word) > 2})
        order_value = row.get("Orden")
        order = int(order_value) if order_value not in (None, "") else int(previous.get("orden") or len(tools) + 1)
        tools.append({
            "id": item_id, "categoriaId": cat_id, "categoria": row["Categoria"], "categoriaIcono": previous.get("categoriaIcono") or old_categories_by_id.get(cat_id, {}).get("icono") or CATEGORY_META[cat_id][1],
            "grupo": row["Grupo"], "icono": row["Icono"], "nombre": row["Nombre"], "tipo": str(row["Tipo"]).casefold(),
            "url": row.get("URL", ""), "webUrl": row.get("WebURL", ""), "package": row.get("Package", ""), "playStore": row.get("PlayStore", ""),
            "notas": row.get("Notas", ""), "favorito": truthy(row.get("Favorito", "")), "orden": order, "keywords": keywords, "estado": "activo"
        })
    tools.sort(key=lambda x: (x["orden"], x["nombre"].casefold()))
    favorites = [t["id"] for t in tools if t["favorito"]]
    categories = []
    for cat_id in dict.fromkeys(t["categoriaId"] for t in tools):
        previous_category = old_categories_by_id.get(cat_id, {})
        name, icon, color = CATEGORY_META[cat_id]
        categories.append({
            "id": cat_id,
            "nombre": previous_category.get("nombre", name),
            "icono": previous_category.get("icono", icon),
            "color": previous_category.get("color", color),
            "descripcion": previous_category.get("descripcion", f"Herramientas de {name.casefold()}."),
            "contador": sum(t["categoriaId"] == cat_id for t in tools),
            "accent": previous_category.get("accent", color),
        })

    identity_rows = {r["Identificador"]: r for r in raw["Identidad"] if truthy(r["Visible"])}
    old_identity = json.loads((data / "identity.json").read_text(encoding="utf-8")) if (data / "identity.json").exists() else {}
    identity = {
        "hero": {
            **old_identity.get("hero", {}),
            "greeting": {
                "morning": identity_rows.get("hero.greeting.morning", {}).get("Valor", "Buenos días Partners."),
                "afternoon": identity_rows.get("hero.greeting.afternoon", {}).get("Valor", "Buenas tardes Partners."),
                "evening": identity_rows.get("hero.greeting.evening", {}).get("Valor", "Buenas noches Partners."),
            },
            "campaign": {
                "primary": identity_rows.get("hero.campaign.primary", {}).get("Valor", ""),
                "accent": identity_rows.get("hero.campaign.accent", {}).get("Valor", ""),
                "display": identity_rows.get("hero.campaign.display", {}).get("Valor", ""),
                "primaryColor": identity_rows.get("hero.campaign.primary", {}).get("Color", "#006241"),
                "accentColor": identity_rows.get("hero.campaign.accent", {}).get("Color", "#111111"),
                "style": identity_rows.get("hero.campaign.display", {}).get("Estilo", "corporativo-expresivo"),
                "featured": True,
            },
            "hashtags": old_identity.get("hero", {}).get("hashtags", ["#GreenApronService", "#DistritoKike"]),
        }
    }

    bt, ss, tbw = raw["BT"], raw["SS"], raw["TBW"]
    partner_development, partner_report = build_partner_development(bt, ss, tbw)
    celebrations = []
    for row in raw["Aniversarios_Cumpleanos"]:
        if not truthy(row.get("Publicar")):
            continue
        celebrations.append({
            "ID": row.get("ID", ""), "Tipo": row.get("Tipo", ""),
            "NOMBRE": row.get("Nombre Partner", ""), "Fecha": row.get("Fecha", ""),
            "PUESTO": row.get("Puesto", ""), "TIENDA": row.get("Tienda", ""),
            "DM": row.get("Distrito", ""), "REGION": row.get("Región", ""),
            "Publicar": True, "NUM_EMP": row.get("NUM_EMP", ""), "CECO": row.get("CECO", "")
        })
    operational = {
        "eventos": events, "actividadesDiarias": daily, "actividadesSemanales": weekly,
        "dutyRoster": duty_roster, "dutyDetail": duty_detail, "checklistApertura": checklist,
        "altasCurso": {"bt": bt, "ss": ss, "tbw": tbw}, "wfm": raw["WFM"],
        "cmsFuente": "Distrito_Go_CMS_v2_actualizado.xlsx", "informativo": info, "celebraciones": celebrations,
        "wfmRegla": raw["WFM"][0].get("Regla WFM", "") if raw["WFM"] else "",
    }

    outputs = {
        "actividades-diarias.v10.json": daily, "actividades-semanales.v10.json": weekly,
        "informativo.v10.json": info, "eventos.v10.json": events,
        "duty-roster.v10.json": duty_roster, "duty-detail.v10.json": duty_detail,
        "checklist-apertura.v10.json": checklist, "bt.json": bt, "ss.json": ss, "tbw.json": tbw,
        "wfm.json": raw["WFM"], "links.json": links, "herramientas.v10.json": tools,
        "favoritos.v10.json": favorites, "categorias.v10.json": categories,
        "altas-curso.v10.json": {"bt": bt, "ss": ss, "tbw": tbw}, "identity.json": identity,
        "desarrollo-partner.v1.json": partner_development,
        "operacional.v10.json": operational,
        "cms-build.v1.json": source_metadata(root, sheets, source_path),
    }
    for name, value in outputs.items():
        emit(name, value)
    emit_report("desarrollo-partner-build.json", partner_report)
    return changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("cms", type=Path)
    parser.add_argument("--project", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()
    sheets, errors = validate_workbook(args.cms)
    sheets = apply_header_overlays(args.project.resolve(), sheets)
    if errors:
        print("\n".join(f"ERROR: {e}" for e in errors))
        return 1
    print("CMS válido:", ", ".join(f"{k}={len(v)}" for k, v in sheets.items()))
    if not args.validate_only:
        changed = build(args.project.resolve(), sheets, args.cms.resolve())
        print(f"JSON actualizados: {len(changed)}")
        for path in changed:
            print(path.relative_to(args.project.resolve()).as_posix())
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
