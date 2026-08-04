#!/usr/bin/env python3
"""Valida estructura, calidad de datos, fechas y recursos del CMS."""
from __future__ import annotations

import argparse
import calendar
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from cms_pipeline import validate_workbook

TRUE_VALUES = {"true", "verdadero", "si", "sí", "1", "yes"}
FALSE_VALUES = {"false", "falso", "no", "0"}
BOOLEAN_VALUES = TRUE_VALUES | FALSE_VALUES
LOCAL_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".avif", ".gif", ".svg"}
EVENT_ACTIONS = {"enlace", "imagen", "informativo"}


def text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date()
    except ValueError:
        return None


def local_asset_exists(cms: Path, value: Any) -> bool:
    raw = text(value).split("?", 1)[0].split("#", 1)[0]
    if not raw or re.match(r"^https?://", raw, re.I):
        return True
    name = Path(raw).name
    if Path(name).suffix.casefold() not in LOCAL_IMAGE_EXTENSIONS:
        return True
    root = cms.resolve().parent
    candidates = [root / raw, root / "assets" / "photos" / name, root / "assets" / "img" / name]
    return any(path.is_file() for path in candidates)


def complete_http_url(value: Any) -> bool:
    raw = text(value)
    if not re.match(r"^https?://", raw, re.I) or "..." in raw or re.search(r"[{}<>]", raw):
        return False
    return bool(re.match(r"^https?://[^/\s]+(?:/[^\s]*)?$", raw, re.I))


def duplicate_values(rows: list[dict[str, Any]], column: str) -> list[str]:
    seen: set[str] = set()
    duplicated: set[str] = set()
    for row in rows:
        value = text(row.get(column)).casefold()
        if not value:
            continue
        if value in seen:
            duplicated.add(text(row.get(column)))
        seen.add(value)
    return sorted(duplicated, key=str.casefold)


def formula_cells_without_cached_value(cms: Path) -> list[str]:
    formulas = load_workbook(cms, read_only=True, data_only=False)
    values = load_workbook(cms, read_only=True, data_only=True)
    missing: list[str] = []
    for sheet_name in formulas.sheetnames:
        formula_sheet = formulas[sheet_name]
        value_sheet = values[sheet_name]
        for row in formula_sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and value_sheet[cell.coordinate].value is None:
                    missing.append(f"{sheet_name}!{cell.coordinate}")
    formulas.close()
    values.close()
    return missing


def audit_cms(cms: Path) -> dict[str, Any]:
    sheets, structural_errors = validate_workbook(cms)
    critical = list(structural_errors)
    warnings: list[str] = []
    metrics: dict[str, Any] = {"sheets": {name: len(rows) for name, rows in sheets.items()}}

    contract_path = cms.resolve().parent / "cms-contract.json"
    contract = json.loads(contract_path.read_text(encoding="utf-8")) if contract_path.is_file() else {"sheets": {}}
    for sheet_name, spec in contract.get("sheets", {}).items():
        primary = spec.get("primaryKey")
        if not primary:
            continue
        columns = [primary] if isinstance(primary, str) else list(primary)
        seen: set[tuple[str, ...]] = set()
        for row in sheets.get(sheet_name, []):
            row_number = row.get("__row__", "?")
            key = tuple(text(row.get(column)) for column in columns)
            if not all(key):
                critical.append(f"{sheet_name} fila {row_number}: clave obligatoria incompleta ({', '.join(columns)})")
            elif key in seen:
                critical.append(f"{sheet_name} fila {row_number}: clave duplicada {' / '.join(key)}")
            seen.add(key)

    uncached_formulas = formula_cells_without_cached_value(cms)
    if uncached_formulas:
        critical.append("Fórmulas sin resultado guardado: " + ", ".join(uncached_formulas[:20]))

    events = sheets.get("Eventos", [])
    inventory = {"weekly": 0, "monthEnd": 0}
    for row in events:
        row_number = row.get("__row__", "?")
        event_id = text(row.get("ID")) or f"fila {row_number}"
        name = text(row.get("Nombre Evento"))
        start = as_date(row.get("Fecha Inicio"))
        end = as_date(row.get("Fecha Fin"))
        publish = text(row.get("Publicar")).casefold()
        action = text(row.get("Tipo de acción")).casefold()
        if not name:
            critical.append(f"Eventos {event_id}: falta Nombre Evento")
        if not start or not end:
            critical.append(f"Eventos {event_id}: Fecha Inicio y Fecha Fin deben ser fechas válidas")
            continue
        if start > end:
            critical.append(f"Eventos {event_id}: Fecha Inicio es posterior a Fecha Fin")
        if publish not in BOOLEAN_VALUES:
            critical.append(f"Eventos {event_id}: Publicar debe ser VERDADERO/FALSO o Sí/No")
        if action not in EVENT_ACTIONS:
            critical.append(f"Eventos {event_id}: Tipo de acción debe ser Enlace, Imagen o Informativo")
        normalized = name.casefold()
        if normalized == "inventario semanal":
            inventory["weekly"] += 1
            if start != end or start.weekday() != 6:
                critical.append(f"Eventos {event_id}: Inventario semanal debe iniciar y terminar el mismo domingo")
        elif normalized == "inventario fin de mes":
            inventory["monthEnd"] += 1
            last_day = calendar.monthrange(start.year, start.month)[1]
            if start != end or start.day != last_day:
                critical.append(f"Eventos {event_id}: Inventario fin de mes debe iniciar y terminar el último día calendario")
        resource = row.get("Link/Imagen")
        image_resource = row.get("Imagen")
        if resource and re.match(r"^https?://", text(resource), re.I) and not complete_http_url(resource):
            critical.append(f"Eventos {event_id}: enlace incompleto o inválido {text(resource)}")
        if action == "enlace" and not complete_http_url(resource):
            critical.append(f"Eventos {event_id}: Tipo de acción Enlace requiere una URL HTTPS completa")
        if action == "imagen" and not any(
            Path(text(value)).suffix.casefold() in LOCAL_IMAGE_EXTENSIONS for value in (resource, image_resource) if text(value)
        ):
            critical.append(f"Eventos {event_id}: Tipo de acción Imagen requiere un archivo gráfico")
        if resource and not local_asset_exists(cms, resource):
            critical.append(f"Eventos {event_id}: no existe el recurso local {text(resource)}")

    metrics["inventoryEvents"] = inventory

    info = sheets.get("Informativo", [])
    visible_home = 0
    for row in info:
        row_number = row.get("__row__", "?")
        item_id = text(row.get("ID")) or f"fila {row_number}"
        if not text(row.get("Actividad")):
            critical.append(f"Informativo {item_id}: falta Actividad")
        visible = text(row.get("Visible")).casefold()
        if visible not in BOOLEAN_VALUES:
            critical.append(f"Informativo {item_id}: Visible debe ser VERDADERO/FALSO o Sí/No")
        resource = row.get("Link /Imagen")
        if not text(resource):
            warnings.append(f"Informativo {item_id}: no tiene Link /Imagen")
        elif not local_asset_exists(cms, resource):
            critical.append(f"Informativo {item_id}: no existe el recurso local {text(resource)}")
        if text(row.get("Mostrar Inicio")).casefold() in TRUE_VALUES:
            visible_home += 1
        valid_from = as_date(row.get("Vigencia Inicio"))
        valid_to = as_date(row.get("Vigencia Fin"))
        if row.get("Vigencia Inicio") and not valid_from:
            critical.append(f"Informativo {item_id}: Vigencia Inicio no es una fecha válida")
        if row.get("Vigencia Fin") and not valid_to:
            critical.append(f"Informativo {item_id}: Vigencia Fin no es una fecha válida")
        if valid_from and valid_to and valid_from > valid_to:
            critical.append(f"Informativo {item_id}: Vigencia Inicio es posterior a Vigencia Fin")
    if visible_home > 1:
        warnings.append(f"Informativo: {visible_home} registros están marcados para Mostrar Inicio; se recomienda uno")

    for row in sheets.get("Links", []):
        row_number = row.get("__row__", "?")
        name = text(row.get("Nombre")) or f"fila {row_number}"
        destinations = [text(row.get(key)) for key in ("URL", "WebURL", "Package", "PlayStore")]
        if not any(destinations):
            critical.append(f"Links {name}: no tiene ningún destino")
        for value in destinations[:2]:
            if value and not re.match(r"^(?:https?|intent)://", value, re.I):
                critical.append(f"Links {name}: URL inválida {value}")

    metrics["records"] = sum(len(rows) for rows in sheets.values())
    metrics["informativeHomeItems"] = visible_home
    return {
        "ok": not critical,
        "cms": cms.name,
        "summary": {"critical": len(critical), "warnings": len(warnings)},
        "metrics": metrics,
        "errors": critical,
        "warnings": warnings,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("cms", type=Path)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    report = audit_cms(args.cms)
    output = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(output, encoding="utf-8")
    print(output, end="")
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
