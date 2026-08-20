#!/usr/bin/env python3
"""Reconciliación dinámica entre las 14 pestañas, metadatos y JSON publicados."""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from cms_pipeline import build, canonical_hash, truthy, validate_workbook

ROOT = Path(__file__).resolve().parents[1]


def clean_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{key: value for key, value in row.items() if key != "__row__"} for row in rows]


def load(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def compare_current(cms: Path, project: Path) -> dict[str, Any]:
    sheets, errors = validate_workbook(cms)
    if errors:
        return {"ok": False, "errors": errors, "checks": []}
    metadata = load(project / "data" / "cms-build.v1.json")
    checks: list[dict[str, Any]] = []

    def check(name: str, condition: bool, detail: str) -> None:
        checks.append({"name": name, "ok": bool(condition), "detail": detail})

    source_sha = hashlib.sha256(cms.read_bytes()).hexdigest()
    check("Excel identificado por SHA-256", metadata.get("sourceSha256") == source_sha, source_sha)
    check("Total de registros", metadata.get("records") == sum(map(len, sheets.values())), str(metadata.get("records")))
    for name, rows in sheets.items():
        entry = metadata.get("sheets", {}).get(name, {})
        expected_hash = canonical_hash(clean_rows(rows))
        check(f"Pestaña {name}", entry.get("records") == len(rows) and entry.get("contentSha256") == expected_hash, f"{len(rows)} registros")

    operational = load(project / "data" / "operacional.v10.json")
    published_informative = load(project / "data" / "informativo.v10.json")
    mappings = {
        "Informativo": ("data/informativo.v10.json", len(sheets["Informativo"])),
        "WFM": ("data/wfm.json", len(sheets["WFM"])),
        "BT": ("data/bt.json", len(sheets["BT"])),
        "TBW": ("data/tbw.json", len(sheets["TBW"])),
        "SS": ("data/ss.json", len(sheets["SS"])),
        "Links": ("data/links.json", len(sheets["Links"])),
        "Eventos": ("data/eventos.v10.json", sum(truthy(row.get("Publicar")) for row in sheets["Eventos"])),
        "Actividades_Semanales": ("data/actividades-semanales.v10.json", len(sheets["Actividades_Semanales"])),
        "Actividades_Diaria": ("data/actividades-diarias.v10.json", len(sheets["Actividades_Diaria"])),
        "Duty_Roster": ("data/duty-roster.v10.json", len(sheets["Duty_Roster"])),
        "Duty_Detail": ("data/duty-detail.v10.json", len(sheets["Duty_Detail"])),
        "Checklist_Apertura": ("data/checklist-apertura.v10.json", len(sheets["Checklist_Apertura"])),
    }
    for sheet, (relative, expected) in mappings.items():
        actual = len(load(project / relative))
        check(f"Salida {sheet}", actual == expected, f"{actual}/{expected}")

    operational_informative = operational.get("informativo", [])
    check(
        "Informativo integrado en la aplicación",
        operational_informative == published_informative,
        f"{len(operational_informative)}/{len(published_informative)} registros; contenido idéntico",
    )

    celebration_expected = sum(truthy(row.get("Publicar")) for row in sheets["Aniversarios_Cumpleanos"])
    check("Salida Aniversarios_Cumpleanos", len(operational.get("celebraciones", [])) == celebration_expected, f"{len(operational.get('celebraciones', []))}/{celebration_expected}")
    identity_text = json.dumps(load(project / "data" / "identity.json"), ensure_ascii=False)
    visible_identity = [row for row in sheets["Identidad"] if truthy(row.get("Visible"))]
    check("Salida Identidad", all(str(row.get("Valor", "")) in identity_text for row in visible_identity), f"{len(visible_identity)} valores visibles")
    return {"ok": all(item["ok"] for item in checks), "checks": checks, "errors": []}


def mutation_test(cms: Path, project: Path) -> dict[str, Any]:
    """Simula cambiar # Evento y descripción; el resto debe permanecer idéntico."""
    with tempfile.TemporaryDirectory(prefix="distrito-go-cms-") as temp:
        base = Path(temp) / "base"
        changed = Path(temp) / "changed"
        for target in (base, changed):
            target.mkdir()
            shutil.copytree(project / "data", target / "data")
            shutil.copytree(project / "assets", target / "assets")
            (target / "reports").mkdir()
        workbook = Path(temp) / cms.name
        shutil.copy2(cms, workbook)
        wb = load_workbook(workbook)
        ws = wb["Eventos"]
        headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value is not None}
        row_number = next(row for row in range(2, ws.max_row + 1) if ws.cell(row, headers["ID"]).value)
        old_id = str(ws.cell(row_number, headers["ID"]).value)
        new_id = f"{old_id}-QA"
        marker = " [PRUEBA DE SINCRONIZACIÓN]"
        ws.cell(row_number, headers["ID"]).value = new_id
        ws.cell(row_number, headers["Descripción"]).value = str(ws.cell(row_number, headers["Descripción"]).value or "") + marker
        wb.save(workbook)

        original_sheets, original_errors = validate_workbook(cms)
        changed_sheets, changed_errors = validate_workbook(workbook)
        if original_errors or changed_errors:
            return {"ok": False, "detail": "; ".join(original_errors + changed_errors)}
        build(base, original_sheets, cms)
        build(changed, changed_sheets, workbook)
        before = load(base / "data" / "eventos.v10.json")
        after = load(changed / "data" / "eventos.v10.json")
        old = next((item for item in before if str(item.get("ID")) == old_id), None)
        new = next((item for item in after if str(item.get("ID")) == new_id), None)
        untouched_before = {str(item.get("ID")): item for item in before if str(item.get("ID")) != old_id}
        untouched_after = {str(item.get("ID")): item for item in after if str(item.get("ID")) != new_id}
        ok = bool(old and new and marker in str(new.get("Contexto / Recordatorio")) and len(before) == len(after) and untouched_before == untouched_after)
        return {"ok": ok, "detail": f"{old_id} → {new_id}; {len(before)} eventos; resto intacto={untouched_before == untouched_after}"}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("cms", type=Path, nargs="?", default=ROOT / "Distrito_Go_CMS_v2_actualizado.xlsx")
    parser.add_argument("--project", type=Path, default=ROOT)
    parser.add_argument("--mutation-test", action="store_true")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    project = args.project.resolve()
    report = compare_current(args.cms.resolve(), project)
    if args.mutation_test:
        report["mutationTest"] = mutation_test(args.cms.resolve(), project)
        report["ok"] = report["ok"] and report["mutationTest"]["ok"]
    output = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(output, encoding="utf-8")
    print(output, end="")
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
