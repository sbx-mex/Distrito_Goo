#!/usr/bin/env python3
"""Pruebas reproducibles para la búsqueda global v25 de Distrito Goo."""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
CHECKS: list[dict[str, object]] = []


def check(name: str, condition: bool, detail: str) -> None:
    CHECKS.append({"name": name, "ok": bool(condition), "detail": detail})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--report", type=Path, default=ROOT / "reports" / "v25-validation.json")
    args = parser.parse_args()

    html = (ROOT / "index.html").read_text(encoding="utf-8")
    sw = (ROOT / "sw.js").read_text(encoding="utf-8")
    operational = (ROOT / "modules" / "operational.js").read_text(encoding="utf-8")
    experience = (ROOT / "modules" / "experience.js").read_text(encoding="utf-8")
    search = (ROOT / "modules" / "search.js").read_text(encoding="utf-8")
    app = (ROOT / "modules" / "app.js").read_text(encoding="utf-8")
    info = json.loads((ROOT / "data" / "informativo.v10.json").read_text(encoding="utf-8"))

    wb = load_workbook(ROOT / "Distrito_Go_CMS_v2_actualizado.xlsx", read_only=True, data_only=True)
    ws = wb["Informativo"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(value not in (None, "") for value in row)]
    names = {str(row.get("Actividad", "")) for row in records}
    check("CMS visual completo", len(records) == 5, f"{len(records)} registros en Informativo")
    check(
        "Apartados CMS",
        {
            "Maquila",
            "Dress Code Portafolio",
            "Registro Clock In/Out",
            "Coffee Master 2026",
            "Resumen Semanal",
        }.issubset(names),
        ", ".join(sorted(names)),
    )
    check("Resumen semanal CMS", any(row.get("Frecuencia") == "Semanal" for row in records), "Frecuencia Semanal presente")
    visual_headers = {"Etiqueta", "Vigencia Inicio", "Vigencia Fin", "Orden", "Mostrar Inicio", "Mostrar Explorar", "Acceso Rápido"}
    check("Controles visuales CMS", visual_headers.issubset(set(headers)), ", ".join(sorted(visual_headers)))
    check("Prioridad única CMS", sum(str(row.get("Mostrar Inicio", "")).casefold() in {"sí","si","true","1"} for row in records) == 1, "Un registro visible en Inicio")

    check("Contenedor CMS Coffee", 'id="coffee-master-grid"' in html, "Coffee Master sin ruta fija en HTML")
    check("Contenedor CMS semanal", 'id="weekly-updates-grid"' in html, "Actualizaciones sin ruta fija en HTML")
    check(
        "Sin ruta visual fija",
        "data-image-viewer=\"./assets/photos/coffemaster26.jpeg\"" not in html
        and "data-image-viewer=\"./assets/photos/resumen_comunicado_semana_actual.png\"" not in html,
        "Las rutas se reciben del JSON generado",
    )
    check("Navegación primaria app", len(re.findall(r'class="nav-item', html)) == 4, "Inicio, Explorar, Buscar y Guardados")
    check("Navegación agrupada", "navigation-groups" in (ROOT / "modules" / "app.js").read_text(encoding="utf-8"), "Menú Más agrupado")
    check("Carga diferida PDF", "await import('./celebration-pdf.js')" in operational, "PDF se importa al solicitarlo")
    check("Secciones diferidas", "IntersectionObserver" in operational and "data-deferred-section" in html, "Duty y Partner bajo demanda")
    check("Inicio visual", all(token in html for token in ('id="visual-home"', 'id="operational-stories"', 'id="visual-priority-card"', 'id="explore-grid"')), "Historias, prioridad y Explorar presentes")
    check("Cuatro accesos", all(label in experience for label in ("Hoy","Apertura","Personas","Semana")) and "{id:'peak'" not in experience, "Hoy, Apertura, Personas y Semana")
    check("Peak agrupado en Operación", "keywords:['peak','ritmo','cobertura','despliegue','turno']" in experience and "access:'Operación'" in experience, "Duty Roster conserva Peak como palabra clave operativa")
    check("Accesos con CMS", "matchesAccess" in experience and "Acceso Rápido" in experience, "Clasificación compatible con CMS")
    check("Explorar filtrable", "data-explore-category" in experience and "CATEGORIES" in experience, "Filtros sin recarga")
    check("Guardados locales", "dgx_saved_content" in experience and "localStorage" not in experience, "Persistencia mediante abstracción local existente")
    check("Buscador global", all(value in html for value in ("Buscar en Distrito Goo","global-search-results","global-search-status","clear-global-search")) and "createSearchIndex" in search, "Índice único y resultados contiguos")
    check("Jerarquía de búsqueda", html.index('id="global-search-results"') < html.index('class="tools-section"') < html.index('class="categories-section"'), "Resultados, herramientas y categorías en ese orden")
    check("Búsqueda diferida", "const SEARCH_DELAY = 200" in search and "setTimeout(() => renderSearchResults" in search, "Espera de 200 ms")
    check("Búsqueda normalizada", "normalize(query)" in search and "scoreEntry" in search, "Mayúsculas y acentos normalizados")
    check("Estados de búsqueda", all(value in search for value in ("Busca en todo Distrito Goo","Buscando en Distrito Goo…","No encontramos resultados para")), "Inicial, procesando y sin resultados")
    check("Detalle interno", "dgx:open-detail" in experience and "openContentDetail" in app, "Visor existente ampliado")
    check("Portadas estándar", "standardCoverMarkup" in experience and "shouldUseStandardCover" in experience, "Maquila e infografías sin fondo saturado")
    check("Eventos limpios", all(value in html for value in ('data-event-filter="upcoming"', 'data-event-filter="week"', 'data-event-filter="month"')) and "data-open-event-id" in operational, "Próximos, semana y mes")
    check("Sin controles superiores", all(value not in html for value in ("open-spotlight","theme-toggle","spotlight-modal","⌘K")), "Spotlight y tema retirados")
    check("Sin atajos Spotlight", "metaKey" not in search and "ctrlKey" not in search, "Atajos exclusivos retirados")
    check("Pie corporativo", all(text in html for text in ("Diseñado por Enrique César Flores", "#DistritoKike 🚀", "#GreenApronService", "JUNTÉMONOS MÁS")), "Identidad y uso interno presentes una vez")

    shell_match = re.search(r"const APP_SHELL = \[(.*?)\];", sw, re.S)
    shell_refs = re.findall(r"['\"]([^'\"]+)['\"]", shell_match.group(1)) if shell_match else []
    shell_bytes = 0
    for ref in shell_refs:
        path = ROOT / ref.split("?", 1)[0].removeprefix("./")
        if path.is_file():
            shell_bytes += path.stat().st_size
    check("APP_SHELL esencial", shell_bytes < 2_000_000, f"{shell_bytes} bytes")
    check("Sin imágenes pesadas precargadas", not any("/photos/" in ref or "/duty-roster/" in ref for ref in shell_refs), f"{len(shell_refs)} recursos esenciales")
    check("Experiencia offline", "./styles/experience.css" in shell_refs and "./modules/experience.js" in shell_refs, "CSS y módulo visual en APP_SHELL")
    check(
        "Resumen network-first",
        "resumen_comunicado_semana_actual.png" in sw and "event.respondWith(networkFirst(request))" in sw,
        "Excepción exclusiva conservada",
    )

    versioned = [item.get("Recurso", "") for item in info if item.get("TipoRecurso") == "imagen"]
    check("Rutas versionadas", all("?v=" in value for value in versioned), f"{len(versioned)} rutas con hash")
    weekly = next((item for item in info if item.get("Frecuencia") == "Semanal"), {})
    check("Nombre semanal conservado", "resumen_comunicado_semana_actual.png?v=" in weekly.get("Recurso", ""), weekly.get("Recurso", ""))

    webps = sorted((ROOT / "assets").rglob("*.webp"))
    valid_webps = 0
    for path in webps:
        try:
            with Image.open(path) as image:
                image.verify()
            valid_webps += 1
        except Exception:
            pass
    check("WebP válidos", valid_webps == len(webps) and len(webps) >= 8, f"{valid_webps}/{len(webps)} válidos")
    thumbnails = [path for path in webps if path.name.endswith(".thumb.webp")]
    check("Miniaturas WebP", len(thumbnails) >= 8, f"{len(thumbnails)} miniaturas derivadas")
    check("Picture con respaldo", "<picture>" in experience and "imageOriginal" in experience, "WebP con original preservado")
    check("Carga diferida", 'loading="${eager ? \'eager\' : \'lazy\'}"' in experience and 'decoding="async"' in experience, "Lazy fuera del primer bloque")
    check("Caché actualizado", "distrito-go-v25.0.0-busqueda-global" in sw and "./styles/clean.css" in shell_refs, "Versión v25")
    check("Caché tolera versiones", "ignoreSearch:true" in sw, "CSS y JS versionados disponibles offline")

    cms_workflow = ROOT / ".github" / "workflows" / "actualizar-cms.yml"
    check("Workflow CMS", cms_workflow.is_file(), cms_workflow.as_posix())
    check("Workflow temporal retirado", not (ROOT / ".github" / "workflows" / "limpieza-auditada.yml").exists(), "La limpieza auditada ya terminó")
    check("CMS exacto en workflow", "Distrito_Go_CMS_v2_actualizado.xlsx" in cms_workflow.read_text(encoding="utf-8"), "Nombre raíz validado")

    report = {"ok": all(item["ok"] for item in CHECKS), "checks": CHECKS}
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
